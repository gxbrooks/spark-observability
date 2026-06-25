#!/usr/bin/env python3
"""Build multi-sheet Excel workbook comparing ServiceNow and Dynatrace model exports."""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path


def field_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        return str(value.get("display_value") or value.get("value") or "")
    return str(value)


def normalize_host(name) -> str:
    name = field_value(name)
    if not name:
        return ""
    base = name.split(".")[0]
    return base.lower()


def load_data(path: Path) -> dict:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def iter_dt_entities(entities: dict, entity_type: str) -> list[dict]:
    rows: list[dict] = []
    for key, items in (entities or {}).items():
        if key != entity_type and not key.endswith(f"::{entity_type}"):
            continue
        for ent in items or []:
            rows.append(ent)
    return rows


def format_management_zones(ent: dict) -> str:
    mzs = ent.get("managementZones") or ent.get("management_zones") or []
    if not mzs:
        return ""
    if isinstance(mzs, str):
        return mzs
    if isinstance(mzs[0], dict):
        return ", ".join(m.get("name", "") for m in mzs if m.get("name"))
    return ", ".join(str(m) for m in mzs if m)


def flatten_dt_hosts(entities: dict) -> list[dict]:
    rows: list[dict] = []
    for ent in iter_dt_entities(entities, "HOST"):
        rows.append(
            {
                "management_zones": format_management_zones(ent),
                "display_name": ent.get("displayName", ""),
                "entity_id": ent.get("entityId", ""),
                "discovered_name": ent.get("discoveredName", ""),
                "normalized_name": normalize_host(ent.get("displayName", "")),
            }
        )
    return rows


def flatten_dt_by_type(entities: dict, suffix: str) -> list[dict]:
    rows: list[dict] = []
    for ent in iter_dt_entities(entities, suffix):
        rows.append(
            {
                "management_zones": format_management_zones(ent),
                "display_name": ent.get("displayName", ""),
                "entity_id": ent.get("entityId", ""),
            }
        )
    return rows


def sn_location(host: dict) -> str:
    loc = host.get("location") or {}
    if isinstance(loc, dict):
        return str(loc.get("display_value") or loc.get("value") or "")
    return str(loc) if loc else ""


def sn_host_rows(hosts: list[dict]) -> list[dict]:
    rows = []
    for host in hosts or []:
        rows.append(
            {
                "name": field_value(host.get("name")),
                "host_name": field_value(host.get("host_name")),
                "location": sn_location(host),
                "sys_id": field_value(host.get("sys_id")),
                "discovery_source": field_value(host.get("discovery_source")),
                "operational_status": field_value(host.get("operational_status")),
            }
        )
    return rows


CANONICAL_TAG_KEY = "servicenow.io/application-service-identifier"
ALTERNATE_TAG_KEYS = frozenset({"app.kubernetes.io/name", "app"})


def tag_binding_counts(tag_bindings: list[dict]) -> tuple[dict[str, int], dict[str, int]]:
    canonical: dict[str, int] = defaultdict(int)
    alternate: dict[str, int] = defaultdict(int)
    for row in tag_bindings or []:
        value = row.get("value") or row.get("val") or ""
        key = row.get("key") or ""
        if not value:
            continue
        if key == CANONICAL_TAG_KEY:
            canonical[str(value)] += 1
        elif key in ALTERNATE_TAG_KEYS:
            alternate[str(value)] += 1
    return dict(canonical), dict(alternate)


def build_host_diff(sn_hosts: list[dict], dt_hosts: list[dict]) -> tuple[list[dict], list[dict], list[dict]]:
    sn_by_norm = {normalize_host(h.get("name", "")): h for h in sn_hosts or [] if field_value(h.get("name"))}
    dt_by_norm: dict[str, dict] = {}
    for row in dt_hosts:
        norm = row.get("normalized_name") or normalize_host(row.get("display_name", ""))
        if norm and norm not in dt_by_norm:
            dt_by_norm[norm] = row

    matched, sn_only, dt_only = [], [], []
    for norm, sn in sorted(sn_by_norm.items()):
        dt = dt_by_norm.get(norm)
        if dt:
            matched.append(
                {
                    "normalized_name": norm,
                    "servicenow_name": field_value(sn.get("name", "")),
                    "servicenow_sys_id": sn.get("sys_id", ""),
                    "servicenow_location": sn_location(sn),
                    "dynatrace_display_name": dt.get("display_name", ""),
                    "dynatrace_entity_id": dt.get("entity_id", ""),
                    "dynatrace_management_zones": dt.get("management_zones", ""),
                    "status": "matched",
                }
            )
        else:
            sn_only.append(
                {
                    "normalized_name": norm,
                    "servicenow_name": field_value(sn.get("name", "")),
                    "servicenow_sys_id": sn.get("sys_id", ""),
                    "servicenow_location": sn_location(sn),
                    "status": "servicenow_only",
                }
            )

    for norm, dt in sorted(dt_by_norm.items()):
        if norm not in sn_by_norm:
            dt_only.append(
                {
                    "normalized_name": norm,
                    "dynatrace_display_name": dt.get("display_name", ""),
                    "dynatrace_entity_id": dt.get("entity_id", ""),
                    "dynatrace_management_zones": dt.get("management_zones", ""),
                    "status": "dynatrace_only",
                }
            )

    return matched, sn_only, dt_only


def mz_list(ent: dict) -> list[str]:
    raw = format_management_zones(ent)
    return [part.strip() for part in raw.split(",") if part.strip()]


def build_diagnostics(
    unit: dict,
    intent_apps: list[dict],
    sn_hosts: list[dict],
    dt_entities: dict,
    app_diff: list[dict],
) -> list[dict]:
    correlation = unit.get("dynatrace_correlation", {})
    scope = unit.get("scope_unit", {})
    partitioning = correlation.get("partitioning", {})
    expected_mz = partitioning.get("management_zone", "")
    expected_location = scope.get("cmdb_location", "")
    project_hosts = {normalize_host(h) for h in correlation.get("project_host_names", []) if h}
    expected_clusters = {
        c.get("dynatrace_name", ""): c for c in correlation.get("kubernetes_clusters", []) if c.get("dynatrace_name")
    }
    recs = correlation.get("recommendations", {})
    rows: list[dict] = []

    def add_row(**kwargs) -> None:
        rows.append(kwargs)

    # Kubernetes cluster partitioning
    for cluster_name, cluster_meta in sorted(expected_clusters.items()):
        match = next(
            (ent for ent in iter_dt_entities(dt_entities, "KUBERNETES_CLUSTER") if ent.get("displayName") == cluster_name),
            None,
        )
        if not match:
            add_row(
                category="dynatrace_partitioning",
                status="missing_entity",
                entity_type="KUBERNETES_CLUSTER",
                entity_name=cluster_name,
                entity_id="",
                issue="cluster_not_in_tenant",
                detail=f"Expected cluster '{cluster_name}' (SN: {cluster_meta.get('servicenow_name', '')}) not found in Dynatrace export",
                recommendation="Verify DynaKube automatic Kubernetes API monitoring and cluster display name",
            )
            continue
        mzs = mz_list(match)
        if expected_mz and expected_mz not in mzs:
            add_row(
                category="dynatrace_partitioning",
                status="action_required",
                entity_type="KUBERNETES_CLUSTER",
                entity_name=cluster_name,
                entity_id=match.get("entityId", ""),
                issue="missing_management_zone",
                detail=f"Cluster has management zones [{', '.join(mzs) or '(none)'}]; expected '{expected_mz}'",
                recommendation=recs.get("missing_cluster_management_zone", recs.get("missing_management_zone", "")),
            )
        else:
            add_row(
                category="dynatrace_partitioning",
                status="ok",
                entity_type="KUBERNETES_CLUSTER",
                entity_name=cluster_name,
                entity_id=match.get("entityId", ""),
                issue="",
                detail=f"Management zone '{expected_mz}' present" if expected_mz else "Cluster present",
                recommendation="",
            )

    # Project host partitioning
    dt_hosts_by_norm = {normalize_host(ent.get("displayName", "")): ent for ent in iter_dt_entities(dt_entities, "HOST")}
    for norm in sorted(project_hosts):
        ent = dt_hosts_by_norm.get(norm)
        if not ent:
            add_row(
                category="dynatrace_partitioning",
                status="missing_entity",
                entity_type="HOST",
                entity_name=norm,
                entity_id="",
                issue="project_host_not_in_dynatrace",
                detail=f"Project host '{norm}' not found in Dynatrace HOST export",
                recommendation="Verify OneAgent deployment and host group assignment",
            )
            continue
        mzs = mz_list(ent)
        if expected_mz and expected_mz not in mzs:
            add_row(
                category="dynatrace_partitioning",
                status="action_required",
                entity_type="HOST",
                entity_name=ent.get("displayName", norm),
                entity_id=ent.get("entityId", ""),
                issue="missing_management_zone",
                detail=f"Host has management zones [{', '.join(mzs) or '(none)'}]; expected '{expected_mz}'",
                recommendation=recs.get("missing_management_zone", ""),
            )
        else:
            add_row(
                category="dynatrace_partitioning",
                status="ok",
                entity_type="HOST",
                entity_name=ent.get("displayName", norm),
                entity_id=ent.get("entityId", ""),
                issue="",
                detail=f"Management zone '{expected_mz}' present" if expected_mz else "Host present",
                recommendation="",
            )

    # Process group propagation summary
    if expected_mz and project_hosts:
        project_host_mz_ok = sum(
            1 for norm in project_hosts if expected_mz in mz_list(dt_hosts_by_norm.get(norm, {}))
        )
        pg_in_mz = sum(1 for ent in iter_dt_entities(dt_entities, "PROCESS_GROUP") if expected_mz in mz_list(ent))
        pg_total = len(iter_dt_entities(dt_entities, "PROCESS_GROUP"))
        if project_host_mz_ok > 0 and pg_in_mz == 0:
            add_row(
                category="dynatrace_partitioning",
                status="action_required",
                entity_type="PROCESS_GROUP",
                entity_name="(tenant summary)",
                entity_id="",
                issue="missing_process_group_management_zone",
                detail=f"0/{pg_total} process groups in '{expected_mz}' while {project_host_mz_ok}/{len(project_hosts)} project hosts are in that zone",
                recommendation=recs.get("missing_process_group_management_zone", ""),
            )

    # ServiceNow location vs scope hint
    if expected_location:
        for host in sn_hosts or []:
            name = field_value(host.get("name"))
            loc = sn_location(host)
            norm = normalize_host(name)
            if norm not in project_hosts:
                continue
            if loc and loc.lower() != expected_location.lower():
                add_row(
                    category="servicenow_placement",
                    status="action_required",
                    entity_type="cmdb_ci_linux_server",
                    entity_name=name,
                    entity_id=field_value(host.get("sys_id")),
                    issue="location_mismatch",
                    detail=f"CMDB location '{loc}' != scope hint '{expected_location}'",
                    recommendation=recs.get("sn_location_mismatch", ""),
                )

    # Tag-based application services from intent
    for app in app_diff:
        if app.get("status") == "missing_tag_binding":
            add_row(
                category="servicenow_tags",
                status="action_required",
                entity_type="application_service",
                entity_name=app.get("name", ""),
                entity_id=app.get("identifier", ""),
                issue="missing_canonical_tag",
                detail="No cmdb_key_value rows for servicenow.io/application-service-identifier",
                recommendation=recs.get("missing_canonical_tag", ""),
            )
        elif app.get("status") == "ok_alternate_tag_only":
            add_row(
                category="servicenow_tags",
                status="warning",
                entity_type="application_service",
                entity_name=app.get("name", ""),
                entity_id=app.get("identifier", ""),
                issue="alternate_tag_only",
                detail="Only app.kubernetes.io/name or app tag bindings found; canonical servicenow.io key missing",
                recommendation=recs.get("missing_canonical_tag", ""),
            )

    return rows


def build_app_service_diff(
    intent_apps: list[dict],
    sn_apps: list[dict],
    canonical_counts: dict[str, int],
    alternate_counts: dict[str, int],
) -> list[dict]:
    sn_by_name = {a.get("intent_name"): a for a in sn_apps or []}
    rows = []
    for intent in intent_apps or []:
        name = intent.get("name", "")
        identifier = intent.get("identifier", "")
        sn = sn_by_name.get(name, {})
        canonical_count = canonical_counts.get(identifier, 0)
        alternate_count = alternate_counts.get(identifier, 0)
        if not sn.get("present_in_cmdb"):
            status = "missing_in_cmdb"
        elif intent.get("service_mapping") == "tags" and canonical_count > 0:
            status = "ok_canonical_tag"
        elif intent.get("service_mapping") == "tags" and alternate_count > 0:
            status = "ok_alternate_tag_only"
        elif intent.get("service_mapping") == "tags":
            status = "missing_tag_binding"
        elif sn.get("present_in_cmdb"):
            status = "present_in_cmdb"
        else:
            status = "unknown"
        rows.append(
            {
                "name": name,
                "identifier": identifier,
                "platform": intent.get("platform", ""),
                "service_mapping": intent.get("service_mapping", ""),
                "spec_file": intent.get("spec_file", ""),
                "present_in_cmdb": sn.get("present_in_cmdb", False),
                "process_status": sn.get("process_status", ""),
                "service_status": sn.get("service_status", ""),
                "canonical_tag_count": canonical_count,
                "alternate_tag_count": alternate_count,
                "status": status,
            }
        )
    return rows


def cell_value(value) -> str | int | float | bool:
    if value is None:
        return ""
    if isinstance(value, (list, dict, tuple)):
        return json.dumps(value, ensure_ascii=True)
    if isinstance(value, bool):
        return value
    return value


def sheet_from_rows(ws, headers: list[str], rows: list[dict]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([cell_value(row.get(h, "")) for h in headers])


def write_workbook(data: dict, output: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required: pip install openpyxl") from exc

    wb = Workbook()
    # Remove default sheet after we add content
    summary_ws = wb.active
    summary_ws.title = "Summary"

    all_scope_rows = []
    for scope_id, unit in sorted(data.items()):
        scope = unit.get("scope_unit", {})
        intent = unit.get("intent", {})
        sn = unit.get("servicenow", {})
        dt = unit.get("dynatrace", {})
        dt_hosts = flatten_dt_hosts(dt.get("entities", {}))
        matched, sn_only, dt_only = build_host_diff(sn.get("hosts", []), dt_hosts)
        canonical_counts, alternate_counts = tag_binding_counts(sn.get("tag_bindings", []))
        app_diff = build_app_service_diff(
            intent.get("application_services", []),
            sn.get("application_services", []),
            canonical_counts,
            alternate_counts,
        )
        dt_scope_mode = dt.get("scope_mode", "")
        sn_hosts_mode = sn.get("hosts_scope_mode", "")

        all_scope_rows.append(
            {
                "scope_unit_id": scope_id,
                "cmdb_location_hint": scope.get("cmdb_location", ""),
                "cmdb_environment": scope.get("cmdb_environment", ""),
                "dynatrace_mz_hints": ", ".join(scope.get("dynatrace_management_zones", [])),
                "sn_hosts_scope": sn_hosts_mode,
                "dt_entities_scope": dt_scope_mode,
                "intent_app_services": len(intent.get("application_services", [])),
                "sn_hosts": len(sn.get("hosts", [])),
                "sn_app_services_cmdb": len(sn.get("application_services_cmdb", [])),
                "dt_hosts": len(dt_hosts),
                "dt_process_groups": len(flatten_dt_by_type(dt.get("entities", {}), "PROCESS_GROUP")),
                "dt_k8s_clusters": len(flatten_dt_by_type(dt.get("entities", {}), "KUBERNETES_CLUSTER")),
                "hosts_matched": len(matched),
                "hosts_sn_only": len(sn_only),
                "hosts_dt_only": len(dt_only),
                "app_missing_cmdb": sum(1 for r in app_diff if r["status"] == "missing_in_cmdb"),
                "app_missing_tags": sum(1 for r in app_diff if r["status"] == "missing_tag_binding"),
                "app_alternate_tags_only": sum(1 for r in app_diff if r["status"] == "ok_alternate_tag_only"),
                "canonical_tag_bindings": len(sn.get("tag_bindings_canonical", [])),
                "generated_at": unit.get("generated_at", ""),
            }
        )

        safe = re.sub(r"[^\w\-]", "_", scope_id)[:28]

        ws_scope = wb.create_sheet(f"Scope_{safe}"[:31])
        scope_rows = []
        for key, val in sorted(scope.items()):
            scope_rows.append({"field": key, "value": val})
        sheet_from_rows(ws_scope, ["field", "value"], scope_rows)

        ws_intent = wb.create_sheet(f"Intent_{safe}"[:31])
        sheet_from_rows(
            ws_intent,
            ["name", "identifier", "platform", "service_mapping", "parent_business_service", "spec_file", "expand", "host"],
            intent.get("application_services", []),
        )

        ws_sn_hosts = wb.create_sheet(f"SN_Hosts_{safe}"[:31])
        sheet_from_rows(
            ws_sn_hosts,
            ["name", "host_name", "location", "sys_id", "discovery_source", "operational_status"],
            sn_host_rows(sn.get("hosts", [])),
        )

        ws_dt_hosts = wb.create_sheet(f"DT_Hosts_{safe}"[:31])
        sheet_from_rows(ws_dt_hosts, ["management_zones", "display_name", "entity_id", "normalized_name"], dt_hosts)

        ws_host_diff = wb.create_sheet(f"HostDiff_{safe}"[:31])
        sheet_from_rows(
            ws_host_diff,
            [
                "status",
                "normalized_name",
                "servicenow_name",
                "servicenow_location",
                "servicenow_sys_id",
                "dynatrace_display_name",
                "dynatrace_management_zones",
                "dynatrace_entity_id",
            ],
            matched + sn_only + dt_only,
        )

        ws_app = wb.create_sheet(f"AppDiff_{safe}"[:31])
        sheet_from_rows(
            ws_app,
            ["status", "name", "identifier", "platform", "service_mapping", "present_in_cmdb", "canonical_tag_count", "alternate_tag_count", "process_status", "service_status", "spec_file"],
            app_diff,
        )

        ws_tags = wb.create_sheet(f"SN_Tags_{safe}"[:31])
        sheet_from_rows(ws_tags, ["key", "value", "sys_id"], sn.get("tag_bindings", []))

        ws_sn_apps = wb.create_sheet(f"SN_Apps_{safe}"[:31])
        sheet_from_rows(
            ws_sn_apps,
            ["name", "sys_id", "operational_status", "process_status", "service_status"],
            [
                {
                    "name": field_value(row.get("name")),
                    "sys_id": field_value(row.get("sys_id")),
                    "operational_status": field_value(row.get("operational_status")),
                    "process_status": field_value(row.get("process_status")),
                    "service_status": field_value(row.get("service_status")),
                }
                for row in sn.get("application_services_cmdb", [])
            ],
        )

        ws_pg = wb.create_sheet(f"DT_PG_{safe}"[:31])
        sheet_from_rows(
            ws_pg,
            ["management_zones", "display_name", "entity_id"],
            flatten_dt_by_type(dt.get("entities", {}), "PROCESS_GROUP"),
        )

        ws_svc = wb.create_sheet(f"DT_Svc_{safe}"[:31])
        sheet_from_rows(
            ws_svc,
            ["management_zones", "display_name", "entity_id"],
            flatten_dt_by_type(dt.get("entities", {}), "SERVICE"),
        )

        ws_k8s = wb.create_sheet(f"DT_K8s_{safe}"[:31])
        sheet_from_rows(
            ws_k8s,
            ["management_zones", "display_name", "entity_id"],
            flatten_dt_by_type(dt.get("entities", {}), "KUBERNETES_CLUSTER"),
        )

        ws_k8s_node = wb.create_sheet(f"DT_K8sNode_{safe}"[:31])
        sheet_from_rows(
            ws_k8s_node,
            ["management_zones", "display_name", "entity_id"],
            flatten_dt_by_type(dt.get("entities", {}), "KUBERNETES_NODE"),
        )

        diagnostics = build_diagnostics(unit, intent.get("application_services", []), sn.get("hosts", []), dt.get("entities", {}), app_diff)
        ws_diag = wb.create_sheet(f"Diagnostics_{safe}"[:31])
        sheet_from_rows(
            ws_diag,
            ["status", "category", "entity_type", "entity_name", "entity_id", "issue", "detail", "recommendation"],
            diagnostics,
        )

    sheet_from_rows(
        summary_ws,
        list(all_scope_rows[0].keys()) if all_scope_rows else ["scope_unit_id"],
        all_scope_rows,
    )
    summary_ws.insert_rows(1)
    summary_ws["A1"] = "Generated"
    summary_ws["B1"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    wb.save(output)


def main() -> int:
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <compare_data.json> <output.xlsx>", file=sys.stderr)
        return 2
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    data = load_data(input_path)
    write_workbook(data, output_path)
    print(f"Wrote workbook: {output_path}")
    for scope_id, unit in data.items():
        sn = unit.get("servicenow", {})
        dt = unit.get("dynatrace", {})
        dt_hosts = flatten_dt_hosts(dt.get("entities", {}))
        matched, sn_only, dt_only = build_host_diff(sn.get("hosts", []), dt_hosts)
        print(
            f"  {scope_id}: hosts matched={len(matched)} sn_only={len(sn_only)} dt_only={len(dt_only)} "
            f"intent_apps={len(unit.get('intent', {}).get('application_services', []))}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
